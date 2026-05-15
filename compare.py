"""compare.py - 수동 카운트 vs 추론 카운트 비교 엑셀 생성

input/set2/ 영상을 추론하여, manual_count/ 의 사람 카운트와 비교한
정확도 분석 엑셀(xlsx)을 output/ 에 생성한다.

분류 체계는 영상 파일명으로 자동 판별:
  - "12종" 포함 → YOLO custom + Stage-2 vehicle_subtype_12 (12종)
  - 그 외        → YOLO custom only (6종, classifier 없음)
"""

import sys
import os
import time
import hashlib
import json
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent))

import visualize_counting as vc

# ── 경로 설정 (visualize_counting.py 컨벤션) ─────────────────
INPUT_DIR = Path("input/set2")
MANUAL_DIR = Path("manual_count")
OUTPUT_DIR = Path(vc.OUTPUT_DIR)
CONFIG_DIR = Path(vc.CONFIG_DIR)
CACHE_DIR = Path("cache")

MAX_SECONDS = vc.MAX_SECONDS

# ── 6종 비교 카테고리 (manual Excel 표기 → 모델 내부명) ───────
CLS_6_MANUAL_TO_MODEL = {
    "car": "car", "bus-s": "bus_s", "bus-m": "bus_m",
    "truck-s": "truck_s", "truck-m": "truck_m", "truck-x": "truck_x",
}
CLS_6_ORDER = ["car", "bus_s", "bus_m", "truck_s", "truck_m", "truck_x"]
CLS_6_DISPLAY = {
    "car": "car", "bus_s": "bus-s", "bus_m": "bus-m",
    "truck_s": "truck-s", "truck_m": "truck-m", "truck_x": "truck-x",
}

# ── 12종 비교 카테고리 ────────────────────────────────────────
# manual Excel 에서 bus-s / bus-m 가 분리돼 있지만 12종 분류기는 "bus" 하나
# → manual 쪽을 bus 로 합산하여 비교.
CLS_12_MANUAL_TO_MODEL = {
    "car": "car", "bus-s": "bus", "bus-m": "bus",
    "truck_s_a": "truck_s_a", "truck_s_b": "truck_s_b",
    "truck_m_3W": "truck_m_3W", "truck_m_4W": "truck_m_4W", "truck_m_5W": "truck_m_5W",
    "truck_4W_ST": "truck_4W_ST", "truck_4W_FT": "truck_4W_FT",
    "truck_5W_ST": "truck_5W_ST", "truck_5W_FT": "truck_5W_FT",
    "truck_6W_ST": "truck_6W_ST",
}
CLS_12_ORDER = [
    "car", "bus",
    "truck_s_a", "truck_s_b",
    "truck_m_3W", "truck_m_4W", "truck_m_5W",
    "truck_4W_ST", "truck_4W_FT",
    "truck_5W_ST", "truck_5W_FT",
    "truck_6W_ST",
]
CLS_12_DISPLAY = {k: k for k in CLS_12_ORDER}


def is_12class(stem):
    return "12종" in stem


# ═════════════════════════════════════════════════════════════
# manual_count Excel 파싱
# ═════════════════════════════════════════════════════════════

def parse_manual_count(xlsm_path, cls_manual_to_model):
    """manual_count xlsm → {model_cls_name: count}."""
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)
    ws = wb.active
    headers = {}
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and " : " in str(val):
            raw_name = str(val).split(" : ", 1)[1].strip()
            headers[col] = raw_name
    counts = {}
    for col, raw_name in headers.items():
        model_name = cls_manual_to_model.get(raw_name)
        if model_name is None:
            continue
        val = ws.cell(row=11, column=col).value
        v = int(val) if val else 0
        counts[model_name] = counts.get(model_name, 0) + v
    wb.close()
    return counts


# ═════════════════════════════════════════════════════════════
# 추론 결과 캐싱
# ═════════════════════════════════════════════════════════════

def _file_stat(path):
    """파일 경로 + 크기 + 수정시간을 dict로 반환. 없으면 None."""
    p = Path(path)
    if not p.exists():
        return None
    st = p.stat()
    return {"path": str(p), "size": st.st_size, "mtime": st.st_mtime}


def build_fingerprint(video_path, use_stage2):
    """추론 결과에 영향을 주는 모든 파라미터를 계층적 dict로 수집."""
    video_path = Path(video_path)
    cfg_path = CONFIG_DIR / f"{video_path.stem}.json"
    cfg = vc.load_video_config(cfg_path)

    det_cfg = vc.DETECTORS["custom"]
    det_model_path = Path(det_cfg["path"])

    clf_section = None
    if use_stage2:
        clf_key = "vehicle_subtype_12"
        clf_cfg = vc.CLASSIFIERS[clf_key]
        clf_model_path = Path(clf_cfg["path"])
        clf_section = {
            "key": clf_key,
            **(_file_stat(clf_model_path) or {}),
            "smooth_enabled": vc.STAGE2_SMOOTH_ENABLED,
            "smooth_alpha": vc.STAGE2_SMOOTH_ALPHA,
            "min_top1_prob": vc.STAGE2_MIN_TOP1_PROB,
            "min_margin": vc.STAGE2_MIN_MARGIN,
            "bbox_padding": vc.STAGE2_BBOX_PADDING,
            "smooth_ttl_sec": vc.STAGE2_SMOOTH_TTL_SEC,
            "trust_bbox_px": vc.STAGE2_TRUST_BBOX_PX,
            "trust_top1": vc.STAGE2_TRUST_TOP1,
            "trust_alpha": vc.STAGE2_TRUST_ALPHA,
            "trust_adaptive": vc.STAGE2_TRUST_BBOX_ADAPTIVE,
            "trust_pct": vc.STAGE2_TRUST_BBOX_PCT,
            "trust_warmup": vc.STAGE2_TRUST_BBOX_WARMUP,
            "trust_floor": vc.STAGE2_TRUST_BBOX_FLOOR,
            "trust_recompute_every": vc.STAGE2_TRUST_BBOX_RECOMPUTE_EVERY,
            "apply_when": clf_cfg.get("apply_when", "after_tracking"),
            "transform_style": str(clf_cfg.get("transform_style", "timm")),
            "input_size": int(clf_cfg.get("input_size", 300)),
            "min_bbox": int(clf_cfg.get("min_bbox", 32)),
        }

    return {
        "video": _file_stat(video_path),
        "config": {
            "polylines_lst": cfg.get("polylines_lst") if cfg else None,
            "roi_polygon": cfg.get("roi_polygon") if cfg else None,
        },
        "detection": {
            **(_file_stat(det_model_path) or {}),
            "allowed_names": det_cfg.get("allowed_names"),
            "conf": 0.25,
            "iou": 0.45,
            "imgsz": 960,
        },
        "classification": clf_section,
        "counting": {
            "mode": vc.COUNTING_MODE,
            "max_seconds": MAX_SECONDS,
        },
        "roi": {
            "enabled": vc.ROI_ENABLED,
            "outside_behavior": vc.ROI_OUTSIDE_BEHAVIOR,
        },
        "tracker": {
            "use_tracker_bbox": vc.USE_TRACKER_BBOX,
            "use_lost_prediction": vc.TRACKER_USE_LOST_PREDICTION,
            "predict_max_gap": vc.TRACKER_PREDICT_MAX_GAP,
            "track_buffer": 45,
        },
        "hidden_classes": sorted(vc.HIDDEN_CLASSES),
        "center_point_height_offset": vc.CENTER_POINT_HEIGHT_OFFSET,
    }


def fingerprint_hash(fp):
    """fingerprint dict → SHA256 앞 12자."""
    blob = json.dumps(fp, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()[:12]


def load_cache(video_stem, fp_hash):
    """캐시 파일이 존재하면 dict 반환, 없으면 None."""
    path = CACHE_DIR / f"{video_stem}_{fp_hash}.json"
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return None


def save_cache(video_stem, fp_hash, fingerprint, total_class_counts,
               elapsed_sec, frames_processed):
    """추론 결과를 캐시 파일에 저장."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    data = {
        "fingerprint": fingerprint,
        "hash": fp_hash,
        "video_stem": video_stem,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "elapsed_sec": round(elapsed_sec, 2),
        "frames_processed": frames_processed,
        "total_class_counts": total_class_counts,
    }
    path = CACHE_DIR / f"{video_stem}_{fp_hash}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


# ═════════════════════════════════════════════════════════════
# 추론 (process_single_video 카운팅 전용 축소판)
# ═════════════════════════════════════════════════════════════

def run_inference(video_path, use_stage2, yolo_model, yolo_label_mapping,
                  yolo_allowed_classes, tracker_factory, classifier_bundle=None):
    """영상 추론 → (total_class_counts, elapsed_sec, frames_processed) 또는 None."""
    cfg = vc.load_video_config(CONFIG_DIR / f"{video_path.stem}.json")
    if not cfg:
        print(f"  [error] config 없음: {video_path.stem}")
        return None

    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        print(f"  [error] 영상을 열 수 없음: {video_path}")
        return None

    vid_fps = cap.get(cv2.CAP_PROP_FPS)
    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    max_frames = int(MAX_SECONDS * vid_fps) if MAX_SECONDS else total_frames
    max_frames = min(max_frames, total_frames)

    canvas_sz = cfg.get("canvas_size", f"{w}x{h}")
    src_w, src_h = map(int, canvas_sz.split("x"))
    counting_lines = vc.scale_lines(cfg["polylines_lst"], (src_w, src_h), (w, h))

    roi_polygon = None
    if vc.ROI_ENABLED and cfg.get("roi_polygon"):
        roi_raw = np.array(cfg["roi_polygon"], dtype=np.float32)
        if src_w != w or src_h != h:
            roi_raw[:, 0] *= (w / src_w)
            roi_raw[:, 1] *= (h / src_h)
        roi_polygon = roi_raw.astype(np.int32)

    if use_stage2 and classifier_bundle:
        merged = vc.merge_stage1_and_classifier_label_maps(
            yolo_label_mapping,
            classifier_bundle["clf_label_mapping"],
            classifier_bundle["offset"],
        )
        smoother = None
        if vc.STAGE2_SMOOTH_ENABLED:
            num_clf_classes = len(classifier_bundle["clf_label_mapping"])
            ttl_frames = max(1, int(vc.STAGE2_SMOOTH_TTL_SEC * vid_fps))
            smoother = vc.TrackStage2Smoother(
                num_classes=num_clf_classes,
                alpha=float(vc.STAGE2_SMOOTH_ALPHA),
                min_top1=float(vc.STAGE2_MIN_TOP1_PROB),
                min_margin=float(vc.STAGE2_MIN_MARGIN),
                ttl_frames=ttl_frames,
                trust_bbox_px=int(vc.STAGE2_TRUST_BBOX_PX),
                trust_top1=float(vc.STAGE2_TRUST_TOP1),
                trust_alpha=float(vc.STAGE2_TRUST_ALPHA),
                trust_adaptive=bool(vc.STAGE2_TRUST_BBOX_ADAPTIVE),
                trust_pct=float(vc.STAGE2_TRUST_BBOX_PCT),
                trust_warmup=int(vc.STAGE2_TRUST_BBOX_WARMUP),
                trust_floor_px=float(vc.STAGE2_TRUST_BBOX_FLOOR),
                trust_recompute_every=int(vc.STAGE2_TRUST_BBOX_RECOMPUTE_EVERY),
            )
    else:
        merged = yolo_label_mapping
        smoother = None

    tracker = tracker_factory()
    counter = vc.SimpleCounter(counting_lines, merged, target_fps=int(vid_fps))

    from tqdm import tqdm
    t0 = time.perf_counter()
    for frame_idx in tqdm(range(max_frames), total=max_frames,
                          desc=f"  {video_path.stem}", mininterval=0.5):
        ret, frame = cap.read()
        if not ret:
            break
        video_time_ms = int(frame_idx / vid_fps * 1000)

        det = vc._run_yolo_detection(yolo_model, frame, yolo_allowed_classes)

        if (use_stage2 and classifier_bundle
                and classifier_bundle.get("apply_when") == "after_detection"):
            det = vc.apply_stage2_to_detections(
                frame, det, yolo_label_mapping, classifier_bundle,
                smoother=smoother, frame_idx=frame_idx)

        tracked_bboxes = vc._extract_tracked_bboxes(tracker, det, frame)

        if (use_stage2 and classifier_bundle
                and classifier_bundle.get("apply_when") == "after_tracking"):
            tracked_bboxes = vc.apply_stage2_classification(
                frame, tracked_bboxes, yolo_label_mapping, classifier_bundle,
                smoother=smoother, frame_idx=frame_idx)

        tracked_bboxes = vc.filter_boxes_by_roi(tracked_bboxes, merged, roi_polygon)
        counter.update(tracked_bboxes, video_time_ms, frame_idx)

    elapsed = time.perf_counter() - t0
    cap.release()
    print(f"    → {frame_idx + 1} frames, {elapsed:.1f}s ({(frame_idx + 1) / max(elapsed, 0.01):.1f} FPS)")
    return dict(counter.total_class_counts), elapsed, frame_idx + 1


# ═════════════════════════════════════════════════════════════
# Excel 생성
# ═════════════════════════════════════════════════════════════

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

ACCURACY_FORMULA = (
    'IF(AND({m}=0,{i}=0),100,IF({m}=0,0,(1-ABS({i}-{m})/{m})*100))'
)


def _apply_header_style(cell):
    cell.font = HEADER_FONT_WHITE
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN_BORDER


def _apply_cell_style(cell, is_total=False):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if is_total:
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


def _safe_sheet_name(name, max_len=31):
    for ch in "\\/*?:[]":
        name = name.replace(ch, "_")
    return name[:max_len]


def build_detail_sheet(wb, video_stem, cls_order, cls_display,
                       manual_counts, infer_counts):
    """영상별 상세 시트 생성. 반환: (sheet, total_row_number)."""
    ws = wb.create_sheet(title=_safe_sheet_name(video_stem))
    headers = ["차종", "수동 카운트", "추론 카운트", "차이", "정확도(%)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _apply_header_style(cell)

    for ri, cls_key in enumerate(cls_order, 2):
        display = cls_display.get(cls_key, cls_key)
        mc = manual_counts.get(cls_key, 0)
        ic = infer_counts.get(cls_key, 0)

        ws.cell(row=ri, column=1, value=display)
        ws.cell(row=ri, column=2, value=mc)
        ws.cell(row=ri, column=3, value=ic)
        ws.cell(row=ri, column=4).value = f"=C{ri}-B{ri}"
        m_ref, i_ref = f"B{ri}", f"C{ri}"
        ws.cell(row=ri, column=5).value = f"={ACCURACY_FORMULA.format(m=m_ref, i=i_ref)}"
        ws.cell(row=ri, column=5).number_format = '0.0'
        for ci in range(1, 6):
            _apply_cell_style(ws.cell(row=ri, column=ci))

    total_row = len(cls_order) + 2
    data_start, data_end = 2, total_row - 1
    ws.cell(row=total_row, column=1, value="계")
    ws.cell(row=total_row, column=2).value = f"=SUM(B{data_start}:B{data_end})"
    ws.cell(row=total_row, column=3).value = f"=SUM(C{data_start}:C{data_end})"
    ws.cell(row=total_row, column=4).value = f"=C{total_row}-B{total_row}"
    m_ref, i_ref = f"B{total_row}", f"C{total_row}"
    ws.cell(row=total_row, column=5).value = f"={ACCURACY_FORMULA.format(m=m_ref, i=i_ref)}"
    ws.cell(row=total_row, column=5).number_format = '0.0'
    for ci in range(1, 6):
        _apply_cell_style(ws.cell(row=total_row, column=ci), is_total=True)

    ws.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 14

    return ws, total_row


def build_summary_sheet(wb, results):
    """요약 시트 생성 (첫 번째 시트)."""
    ws = wb.active
    ws.title = "요약"
    headers = ["영상명", "분류체계", "수동(계)", "추론(계)", "차이", "정확도(%)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _apply_header_style(cell)

    for ri, res in enumerate(results, 2):
        sheet_name = _safe_sheet_name(res["stem"])
        total_row = res["total_row"]

        ws.cell(row=ri, column=1, value=res["stem"])
        ws.cell(row=ri, column=2, value=res["scheme"])
        ws.cell(row=ri, column=3).value = f"='{sheet_name}'!B{total_row}"
        ws.cell(row=ri, column=4).value = f"='{sheet_name}'!C{total_row}"
        ws.cell(row=ri, column=5).value = f"=D{ri}-C{ri}"
        m_ref, i_ref = f"C{ri}", f"D{ri}"
        ws.cell(row=ri, column=6).value = f"={ACCURACY_FORMULA.format(m=m_ref, i=i_ref)}"
        ws.cell(row=ri, column=6).number_format = '0.0'
        for ci in range(1, 7):
            _apply_cell_style(ws.cell(row=ri, column=ci))

    avg_row = len(results) + 2
    data_start, data_end = 2, avg_row - 1
    ws.cell(row=avg_row, column=1, value="전체")
    ws.cell(row=avg_row, column=3).value = f"=SUM(C{data_start}:C{data_end})"
    ws.cell(row=avg_row, column=4).value = f"=SUM(D{data_start}:D{data_end})"
    ws.cell(row=avg_row, column=5).value = f"=D{avg_row}-C{avg_row}"
    m_ref, i_ref = f"C{avg_row}", f"D{avg_row}"
    ws.cell(row=avg_row, column=6).value = f"={ACCURACY_FORMULA.format(m=m_ref, i=i_ref)}"
    ws.cell(row=avg_row, column=6).number_format = '0.0'
    for ci in range(1, 7):
        _apply_cell_style(ws.cell(row=avg_row, column=ci), is_total=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    for col_letter in ["C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 14

    return ws


# ═════════════════════════════════════════════════════════════
# main
# ═════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  수동 카운트 vs 추론 비교")
    print("=" * 60)

    video_paths = sorted(INPUT_DIR.glob("*.mp4"))
    video_paths = [v for v in video_paths if (MANUAL_DIR / f"{v.stem}.xlsm").exists()]
    if not video_paths:
        print("  [error] 매칭되는 영상-manual_count 쌍이 없습니다.")
        return

    print(f"  대상 영상 {len(video_paths)}개:")
    for vp in video_paths:
        tag = "12종" if is_12class(vp.stem) else "6종"
        print(f"    [{tag}] {vp.name}")

    # ── 캐시 조회 (모델 로드 전에 먼저 확인) ──
    cache_hits = {}
    cache_misses = []
    fp_map = {}
    for vp in video_paths:
        use_12 = is_12class(vp.stem)
        fp = build_fingerprint(vp, use_12)
        fp_hash = fingerprint_hash(fp)
        fp_map[vp.stem] = (fp, fp_hash)
        cached = load_cache(vp.stem, fp_hash)
        if cached:
            cache_hits[vp.stem] = cached
        else:
            cache_misses.append(vp)

    if cache_hits:
        print(f"\n  [cache] 캐시 히트 {len(cache_hits)}건:")
        for stem in cache_hits:
            print(f"    ✓ {stem}")
    if cache_misses:
        print(f"  [cache] 캐시 미스 {len(cache_misses)}건 → 추론 필요:")
        for vp in cache_misses:
            print(f"    ✗ {vp.stem}")

    # ── 모델 로드 (캐시 미스가 있을 때만) ──
    yolo_model = None
    yolo_label_mapping = None
    yolo_allowed_classes = None
    tracker_factory = None
    classifier_bundle = None

    if cache_misses:
        print("\n  모델 로드 중...")
        tracker_factory, device = vc._make_tracker_factory()

        det_cfg = vc.DETECTORS["custom"]
        yolo_model, _ = vc._load_yolo_detection_model(det_cfg["path"])
        full_label_mapping = vc.load_class_info(det_cfg["info"])
        allowed_ids, filtered_mapping = vc.derive_allowed_class_ids(
            full_label_mapping, det_cfg.get("allowed_names"))
        yolo_label_mapping = filtered_mapping if filtered_mapping else full_label_mapping
        yolo_allowed_classes = allowed_ids

        has_12_miss = any(is_12class(vp.stem) for vp in cache_misses)
        if has_12_miss:
            classifier_bundle = vc.load_classifier_bundle("vehicle_subtype_12", device)
        print("  모델 로드 완료.\n")
    else:
        print("\n  [cache] 전체 캐시 히트 — 모델 로드 생략\n")

    # ── 영상별 결과 수집 + 엑셀 생성 ──
    wb = openpyxl.Workbook()
    results = []

    for vp in video_paths:
        stem = vp.stem
        use_12 = is_12class(stem)
        scheme = "12종" if use_12 else "6종"
        cls_map = CLS_12_MANUAL_TO_MODEL if use_12 else CLS_6_MANUAL_TO_MODEL
        cls_order = CLS_12_ORDER if use_12 else CLS_6_ORDER
        cls_display = CLS_12_DISPLAY if use_12 else CLS_6_DISPLAY

        print(f"  [{scheme}] {stem}")
        manual = parse_manual_count(MANUAL_DIR / f"{stem}.xlsm", cls_map)

        if stem in cache_hits:
            infer_raw = cache_hits[stem]["total_class_counts"]
            print(f"    [cache] 캐시에서 로드 (hash={fp_map[stem][1]})")
        else:
            result = run_inference(
                vp, use_stage2=use_12, yolo_model=yolo_model,
                yolo_label_mapping=yolo_label_mapping,
                yolo_allowed_classes=yolo_allowed_classes,
                tracker_factory=tracker_factory,
                classifier_bundle=classifier_bundle if use_12 else None,
            )
            if result is None:
                print(f"    [skip] 추론 실패")
                continue
            infer_raw, elapsed, frames = result
            fp, fp_hash = fp_map[stem]
            save_cache(stem, fp_hash, fp, infer_raw, elapsed, frames)
            print(f"    [cache] 저장 완료 (hash={fp_hash})")

        infer = {k: infer_raw.get(k, 0) for k in cls_order}

        m_total = sum(manual.get(k, 0) for k in cls_order)
        i_total = sum(infer.get(k, 0) for k in cls_order)
        if m_total > 0:
            acc = (1 - abs(i_total - m_total) / m_total) * 100
        else:
            acc = 100.0 if i_total == 0 else 0.0
        print(f"    수동: {m_total}  추론: {i_total}  정확도: {acc:.1f}%")

        _, total_row = build_detail_sheet(
            wb, stem, cls_order, cls_display, manual, infer)
        results.append({
            "stem": stem, "scheme": scheme, "total_row": total_row,
        })

    if not results:
        print("  [error] 처리된 영상이 없습니다.")
        return

    build_summary_sheet(wb, results)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = OUTPUT_DIR / f"compare_result_{timestamp}.xlsx"
    wb.save(str(out_path))
    print(f"\n  출력: {out_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
